"""
按功能分类给工作表着色脚本
Color Worksheets by Function/Data Flow

功能：
- 按数据流阶段给工作表标签着色
- 改善工作表识别度和视觉组织
- 帮助用户快速定位相关工作表

数据流分类：
1. 蓝色 - 主数据层 (Master Data)
2. 绿色 - 订单层 (Order Data)
3. 黄色 - 计算层 (Calculation/Processing)
4. 紫色 - 分析层 (Analytics/Dashboard)
5. 灰色 - 其他/隐藏工作表

作者: Claude Code
创建日期: 2026-01-01
"""

import openpyxl
from pathlib import Path
import logging
from datetime import datetime


# ============================================================================
# 配置和常量
# ============================================================================

class SheetColorConfig:
    """工作表颜色配置 - 按功能分类"""

    # 颜色定义 (RGB Hex)
    COLOR_MASTER = '4472C4'         # 蓝色 - 主数据层
    COLOR_ORDER = '70AD47'          # 绿色 - 订单层
    COLOR_CALCULATION = 'FFC000'    # 黄色 - 计算层
    COLOR_ANALYTICS = '7030A0'      # 紫色 - 分析/仪表板层
    COLOR_OTHER = 'A5A5A5'          # 灰色 - 其他/隐藏
    COLOR_SUPPORT = 'ED7D31'        # 橙色 - 支持/辅助

    # 工作表分类映射
    SHEET_CLASSIFICATION = {
        # 蓝色 - 主数据层 (Master Data)
        '00_SKU_Master': COLOR_MASTER,

        # 绿色 - 订单层 (Order Data)
        '01_Cages_Plan': COLOR_ORDER,
        '02_TrayPack_Order': COLOR_ORDER,
        '03_BulkPack_Order': COLOR_ORDER,
        '04_Bagging_Order': COLOR_ORDER,
        '05_Daily_Orders': COLOR_ORDER,
        '06_Resource_Plan': COLOR_ORDER,

        # 黄色 - 计算层 (Calculation/Processing)
        '07_Labor_Calc': COLOR_CALCULATION,
        '08_Chart_Data': COLOR_CALCULATION,

        # 紫色 - 分析/仪表板层 (Analytics/Dashboard)
        '12_Executive_Dash': COLOR_ANALYTICS,
        '13_Progress_Track': COLOR_ANALYTICS,
        '00_Yield_Rates': COLOR_ANALYTICS,

        # 灰色 - 其他/隐藏工作表
        '09_Pallet_Space': COLOR_OTHER,
        '10_Cone_Line': COLOR_OTHER,
        '14_Weekly_Plan': COLOR_OTHER,
        '15_5Day_Forecast': COLOR_OTHER,
    }

    # 中文标签用于日志
    CHINESE_LABELS = {
        COLOR_MASTER: '主数据层',
        COLOR_ORDER: '订单层',
        COLOR_CALCULATION: '计算层',
        COLOR_ANALYTICS: '分析层',
        COLOR_OTHER: '其他/隐藏',
        COLOR_SUPPORT: '支持层',
    }


class ExcelPathConfig:
    """文件路径配置"""

    BASE_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard\data')
    INPUT_FILE = BASE_DIR / 'v39_Normalized_Styled.xlsx'
    OUTPUT_FILE = BASE_DIR / 'v39_Normalized_Colored.xlsx'
    BACKUP_FILE = BASE_DIR / 'v39_Normalized_Styled_backup_before_coloring.xlsx'
    LOG_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard\logs')


# ============================================================================
# 日志设置
# ============================================================================

def setup_logging():
    """配置日志系统"""
    log_dir = ExcelPathConfig.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"coloring_worksheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

logger = setup_logging()


# ============================================================================
# 主工作表着色类
# ============================================================================

class WorksheetColorizer:
    """工作表着色器"""

    def __init__(self, workbook_path):
        """初始化"""
        self.workbook_path = workbook_path
        self.wb = None

    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"Loading workbook: {self.workbook_path}")
        self.wb = openpyxl.load_workbook(self.workbook_path)
        logger.info(f"Loaded {len(self.wb.sheetnames)} worksheets")

    def create_backup(self, backup_path):
        """创建备份"""
        import shutil
        logger.info(f"Creating backup: {backup_path}")
        shutil.copy2(self.workbook_path, backup_path)
        logger.info("Backup created successfully")

    def color_worksheets(self):
        """给工作表着色"""
        logger.info("=" * 60)
        logger.info("Coloring worksheets by data flow stages")
        logger.info("=" * 60)

        # 统计信息
        colored_count = 0
        unclassified = []

        # 遍历所有工作表
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            # 查找该工作表的颜色分类
            if sheet_name in SheetColorConfig.SHEET_CLASSIFICATION:
                color = SheetColorConfig.SHEET_CLASSIFICATION[sheet_name]
                category = SheetColorConfig.CHINESE_LABELS.get(color, 'Unknown')

                # 应用颜色到工作表标签
                ws.sheet_properties.tabColor = color
                colored_count += 1

                logger.info(f"  [{category:8}] {sheet_name:20} => #{color}")
            else:
                # 未分类的工作表
                unclassified.append(sheet_name)
                logger.warning(f"  [未分类] {sheet_name}")

        logger.info("=" * 60)
        logger.info(f"Coloring Summary:")
        logger.info(f"  - Total worksheets: {len(self.wb.sheetnames)}")
        logger.info(f"  - Colored: {colored_count}")
        logger.info(f"  - Unclassified: {len(unclassified)}")

        if unclassified:
            logger.info(f"  - Unclassified worksheets: {', '.join(unclassified)}")

        logger.info("=" * 60)

        return colored_count > 0

    def print_color_legend(self):
        """打印颜色图例"""
        logger.info("")
        logger.info("=" * 60)
        logger.info("Color Legend - Data Flow Stages")
        logger.info("=" * 60)

        color_sections = {
            SheetColorConfig.COLOR_MASTER: ('蓝色 #4472C4', [
                '00_SKU_Master - 主数据、产品信息参考'
            ]),
            SheetColorConfig.COLOR_ORDER: ('绿色 #70AD47', [
                '01_Cages_Plan - 笼子计划',
                '02_TrayPack_Order - 托盘包装订单',
                '03_BulkPack_Order - 散装订单',
                '04_Bagging_Order - 装袋订单',
                '05_Daily_Orders - 每日订单（核心输入）',
                '06_Resource_Plan - 原料/资源计划',
            ]),
            SheetColorConfig.COLOR_CALCULATION: ('黄色 #FFC000', [
                '07_Labor_Calc - 工时计算、劳动成本',
                '08_Chart_Data - 图表数据中间层',
            ]),
            SheetColorConfig.COLOR_ANALYTICS: ('紫色 #7030A0', [
                '00_Yield_Rates - 产率监控（Yield警告）',
                '12_Executive_Dash - 执行仪表板（主要输出）',
                '13_Progress_Track - 进度追踪',
            ]),
            SheetColorConfig.COLOR_OTHER: ('灰色 #A5A5A5', [
                '09_Pallet_Space - 托盘空间',
                '10_Cone_Line - 锥形线',
                '14_Weekly_Plan - 周计划',
                '15_5Day_Forecast - 5日预测',
            ]),
        }

        for color, (label, sheets) in color_sections.items():
            logger.info(f"\n  {label}")
            logger.info(f"  {'-' * 55}")
            for sheet in sheets:
                logger.info(f"    • {sheet}")

        logger.info("\n" + "=" * 60)

    def save_workbook(self):
        """保存工作簿"""
        output_path = ExcelPathConfig.OUTPUT_FILE
        logger.info(f"Saving workbook to: {output_path}")
        self.wb.save(output_path)
        logger.info("Workbook saved successfully")
        return output_path

    def apply_all_coloring(self):
        """应用所有着色 - 主流程"""
        try:
            logger.info("")
            logger.info("=" * 60)
            logger.info("Starting Worksheet Coloring Process")
            logger.info("=" * 60)

            # Step 1: Load workbook
            self.load_workbook()

            # Step 2: Apply coloring
            success = self.color_worksheets()

            # Step 3: Print legend
            self.print_color_legend()

            if success:
                logger.info("=" * 60)
                logger.info("Coloring process completed successfully")
                logger.info("=" * 60)
                return True
            else:
                logger.error("Coloring process failed")
                return False

        except Exception as e:
            logger.error(f"Error during coloring process: {e}", exc_info=True)
            return False


# ============================================================================
# 主执行函数
# ============================================================================

def main():
    """主执行函数"""

    # 检查输入文件是否存在
    if not ExcelPathConfig.INPUT_FILE.exists():
        logger.error(f"Input file not found: {ExcelPathConfig.INPUT_FILE}")
        return False

    # 创建着色器
    colorizer = WorksheetColorizer(ExcelPathConfig.INPUT_FILE)

    # 创建备份
    colorizer.create_backup(ExcelPathConfig.BACKUP_FILE)

    # 应用着色
    success = colorizer.apply_all_coloring()

    if success:
        # 保存文件
        output_file = colorizer.save_workbook()

        logger.info("")
        logger.info("=" * 60)
        logger.info("FINAL SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Input file:  {ExcelPathConfig.INPUT_FILE}")
        logger.info(f"Output file: {output_file}")
        logger.info("")
        logger.info("Worksheets are now colored by data flow stages!")
        logger.info("")
        logger.info("Color mapping:")
        logger.info("  Blue   → Master Data (主数据)")
        logger.info("  Green  → Order Data (订单层)")
        logger.info("  Yellow → Calculation (计算层)")
        logger.info("  Purple → Analytics/Dashboard (分析层)")
        logger.info("  Gray   → Other/Hidden (其他)")
        logger.info("")
        logger.info("=" * 60)

        return True
    else:
        logger.error("Coloring process failed - see log for details")
        return False


# ============================================================================
# 脚本入口
# ============================================================================

if __name__ == "__main__":
    import sys

    try:
        success = main()
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("Process interrupted by user")
        sys.exit(1)

    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        sys.exit(1)
