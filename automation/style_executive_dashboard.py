"""
12_Executive_Dash 现代灰色主题美化脚本
Modern Gray Theme Styling for Executive Dashboard

功能:
- 应用专业的灰色配色方案
- 改进表头和数据区域的视觉层次
- 保持3个部门的视觉区分
- 优化可读性和专业外观

作者: Claude Code
创建日期: 2026-01-01
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from datetime import datetime
import shutil


# ============================================================================
# 配置和常量
# ============================================================================

class StyleConfig:
    """样式配置类 - 定义所有颜色和样式常量"""

    # 颜色定义 (Hex codes)
    COLOR_HEADER_DARK = '343A40'        # 深炭灰 - 部门标题
    COLOR_HEADER_SLATE = '495057'       # 石板灰 - 员工标题
    COLOR_SUBHEADER = 'ADB5BD'          # 中灰 - 副标题
    COLOR_SECTION_LIGHT = 'E9ECEF'      # 浅灰 - Cut-Up/Bagging
    COLOR_SECTION_WHITE = 'F8F9FA'      # 近白 - Tray Pack
    COLOR_SECTION_MEDIUM = 'DEE2E6'     # 中浅灰 - 员工区
    COLOR_BORDER = 'CED4DA'             # 边框灰
    COLOR_TEXT_WHITE = 'FFFFFF'         # 白色文字
    COLOR_TEXT_DARK = '343A40'          # 深色文字
    COLOR_TEXT_BLACK = '000000'         # 黑色文字

    # 字体配置
    FONT_NAME = 'Calibri'
    FONT_SIZE_HEADER = 12
    FONT_SIZE_SUBHEADER = 11
    FONT_SIZE_DATA = 10

    # 行高配置
    ROW_HEIGHT_HEADER = 35
    ROW_HEIGHT_SUBHEADER = 25
    ROW_HEIGHT_DATA = 15


class ExcelPathConfig:
    """文件路径配置"""

    BASE_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard\data')
    INPUT_FILE = BASE_DIR / 'v39_Normalized.xlsx'
    OUTPUT_FILE = BASE_DIR / 'v39_Normalized_Styled.xlsx'
    BACKUP_FILE = BASE_DIR / 'v39_Normalized_backup_before_styling.xlsx'
    LOG_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard\logs')


# ============================================================================
# 日志设置
# ============================================================================

def setup_logging():
    """配置日志系统"""
    log_dir = ExcelPathConfig.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"styling_executive_dash_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

logger = setup_logging()


# ============================================================================
# 样式工厂类
# ============================================================================

class StyleFactory:
    """创建各种样式对象的工厂类"""

    @staticmethod
    def create_header_font(color=StyleConfig.COLOR_TEXT_WHITE):
        """创建表头字体"""
        return Font(
            name=StyleConfig.FONT_NAME,
            size=StyleConfig.FONT_SIZE_HEADER,
            bold=True,
            color=color
        )

    @staticmethod
    def create_subheader_font():
        """创建副表头字体"""
        return Font(
            name=StyleConfig.FONT_NAME,
            size=StyleConfig.FONT_SIZE_SUBHEADER,
            bold=True,
            color=StyleConfig.COLOR_TEXT_DARK
        )

    @staticmethod
    def create_data_font():
        """创建数据字体"""
        return Font(
            name=StyleConfig.FONT_NAME,
            size=StyleConfig.FONT_SIZE_DATA,
            bold=False,
            color=StyleConfig.COLOR_TEXT_BLACK
        )

    @staticmethod
    def create_fill(color):
        """创建填充样式"""
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )

    @staticmethod
    def create_border():
        """创建边框样式"""
        side = Side(style='thin', color=StyleConfig.COLOR_BORDER)
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def create_alignment(horizontal='center', vertical='center'):
        """创建对齐样式"""
        return Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=False
        )


# ============================================================================
# 主样式应用类
# ============================================================================

class ExecutiveDashStyler:
    """12_Executive_Dash 工作表样式应用器"""

    def __init__(self, workbook_path):
        """初始化"""
        self.workbook_path = workbook_path
        self.wb = None
        self.ws = None
        self.style_factory = StyleFactory()

    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"Loading workbook: {self.workbook_path}")
        self.wb = openpyxl.load_workbook(self.workbook_path)
        self.ws = self.wb['12_Executive_Dash']
        logger.info(f"Loaded worksheet: 12_Executive_Dash ({self.ws.max_row} rows x {self.ws.max_column} cols)")

    def create_backup(self, backup_path):
        """创建备份"""
        logger.info(f"Creating backup: {backup_path}")
        shutil.copy2(self.workbook_path, backup_path)
        logger.info("Backup created successfully")

    def apply_header_styling(self):
        """应用表头样式 (Row 2)"""
        logger.info("Applying header styling (Row 2)...")

        # 定义表头区域和颜色
        header_sections = [
            ('B2', 'D2', StyleConfig.COLOR_HEADER_DARK, 'Cut up'),      # Cut-Up
            ('E2', 'G2', StyleConfig.COLOR_HEADER_DARK, 'Tray pack'),   # Tray Pack
            ('H2', 'J2', StyleConfig.COLOR_HEADER_DARK, 'Bagging'),     # Bagging
            ('L2', 'L2', StyleConfig.COLOR_HEADER_SLATE, 'Total Request'),  # Total Request
            ('M2', 'M2', StyleConfig.COLOR_HEADER_SLATE, 'IN'),         # IN
        ]

        # 设置行高
        self.ws.row_dimensions[2].height = StyleConfig.ROW_HEIGHT_HEADER

        # 应用样式到每个表头区域
        for start_cell, end_cell, bg_color, label in header_sections:
            # 合并单元格 (如果需要)
            if start_cell != end_cell:
                merge_range = f"{start_cell}:{end_cell}"
                try:
                    if merge_range not in self.ws.merged_cells:
                        self.ws.merge_cells(merge_range)
                        logger.info(f"  Merged cells: {merge_range}")
                except Exception as e:
                    logger.warning(f"  Could not merge {merge_range}: {e}")

            # 应用样式
            cell = self.ws[start_cell]
            cell.font = self.style_factory.create_header_font()
            cell.fill = self.style_factory.create_fill(bg_color)
            cell.alignment = self.style_factory.create_alignment()
            cell.border = self.style_factory.create_border()

            # 确保单元格值正确
            if cell.value != label:
                cell.value = label

            logger.info(f"  Styled header: {start_cell} = '{label}'")

        logger.info("Header styling completed")

    def apply_subheader_styling(self):
        """应用副表头样式 (Row 3)"""
        logger.info("Applying sub-header styling (Row 3)...")

        # 设置行高
        self.ws.row_dimensions[3].height = StyleConfig.ROW_HEIGHT_SUBHEADER

        # 应用样式到 B3:M3
        for col in range(2, 14):  # B=2 to M=13
            cell = self.ws.cell(3, col)
            cell.font = self.style_factory.create_subheader_font()
            cell.fill = self.style_factory.create_fill(StyleConfig.COLOR_SUBHEADER)
            cell.alignment = self.style_factory.create_alignment()
            cell.border = self.style_factory.create_border()

        logger.info("Sub-header styling completed")

    def apply_data_area_styling(self):
        """应用数据区域样式 (Rows 4-568)"""
        logger.info("Applying data area styling (Rows 4-568)...")

        # 定义列分组和背景色
        column_groups = [
            (2, 4, StyleConfig.COLOR_SECTION_LIGHT),    # B-D: Cut-Up
            (5, 7, StyleConfig.COLOR_SECTION_WHITE),    # E-G: Tray Pack
            (8, 10, StyleConfig.COLOR_SECTION_LIGHT),   # H-J: Bagging
            (11, 13, StyleConfig.COLOR_SECTION_MEDIUM), # K-M: Employee
        ]

        # 应用样式
        total_cells = 0
        for start_col, end_col, bg_color in column_groups:
            col_range = f"{get_column_letter(start_col)}-{get_column_letter(end_col)}"
            logger.info(f"  Styling columns {col_range} with color {bg_color}...")

            for row in range(4, min(self.ws.max_row + 1, 569)):  # 限制到 568 行
                for col in range(start_col, end_col + 1):
                    cell = self.ws.cell(row, col)

                    # 应用样式
                    cell.font = self.style_factory.create_data_font()
                    cell.fill = self.style_factory.create_fill(bg_color)
                    cell.border = self.style_factory.create_border()

                    # 智能对齐 - 数字右对齐，文本左对齐
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = self.style_factory.create_alignment(horizontal='right')
                    else:
                        cell.alignment = self.style_factory.create_alignment(horizontal='left')

                    total_cells += 1

            logger.info(f"  Columns {col_range} styled ({(end_col-start_col+1)*565} cells)")

        logger.info(f"Data area styling completed ({total_cells} cells)")

    def optimize_column_widths(self):
        """优化列宽"""
        logger.info("Optimizing column widths...")

        for col in range(2, 14):  # B to M
            col_letter = get_column_letter(col)

            # 计算最大内容宽度
            max_length = 0
            for row in range(2, min(50, self.ws.max_row + 1)):  # 检查前50行
                cell = self.ws.cell(row, col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            # 设置列宽 (最小12，最大25)
            adjusted_width = min(max(12, max_length + 2), 25)
            self.ws.column_dimensions[col_letter].width = adjusted_width
            logger.info(f"  Column {col_letter}: width = {adjusted_width}")

        logger.info("Column width optimization completed")

    def validate_formulas(self):
        """验证公式完整性"""
        logger.info("Validating formulas...")

        formula_count = 0
        error_count = 0

        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formula_count += 1
                    if cell.value and ('#REF!' in str(cell.value) or '#VALUE!' in str(cell.value)):
                        error_count += 1
                        logger.warning(f"  Error in formula at {cell.coordinate}: {cell.value}")

        logger.info(f"Validation completed: {formula_count} formulas, {error_count} errors")
        return error_count == 0

    def save_workbook(self, output_path):
        """保存工作簿"""
        logger.info(f"Saving workbook to: {output_path}")
        self.wb.save(output_path)
        logger.info("Workbook saved successfully")

    def apply_all_styling(self):
        """应用所有样式 - 主流程"""
        try:
            logger.info("=" * 60)
            logger.info("Starting 12_Executive_Dash Styling Process")
            logger.info("=" * 60)

            # Step 1: Load workbook
            self.load_workbook()

            # Step 2: Apply styling
            self.apply_header_styling()
            self.apply_subheader_styling()
            self.apply_data_area_styling()
            self.optimize_column_widths()

            # Step 3: Validate
            if not self.validate_formulas():
                logger.warning("Formula validation found errors - please review")

            logger.info("=" * 60)
            logger.info("Styling process completed successfully")
            logger.info("=" * 60)

            return True

        except Exception as e:
            logger.error(f"Error during styling process: {e}", exc_info=True)
            return False


# ============================================================================
# 主执行函数
# ============================================================================

def main():
    """主执行函数"""

    # 检查输入文件是否存在
    if not ExcelPathConfig.INPUT_FILE.exists():
        logger.error(f"Input file not found: {ExcelPathConfig.INPUT_FILE}")
        return False

    # 创建样式应用器
    styler = ExecutiveDashStyler(ExcelPathConfig.INPUT_FILE)

    # 创建备份
    styler.create_backup(ExcelPathConfig.BACKUP_FILE)

    # 应用所有样式
    success = styler.apply_all_styling()

    if success:
        # 保存结果
        styler.save_workbook(ExcelPathConfig.OUTPUT_FILE)

        logger.info("")
        logger.info("=" * 60)
        logger.info("SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Input file:  {ExcelPathConfig.INPUT_FILE}")
        logger.info(f"Output file: {ExcelPathConfig.OUTPUT_FILE}")
        logger.info(f"Backup file: {ExcelPathConfig.BACKUP_FILE}")
        logger.info("")
        logger.info("Styling applied successfully!")
        logger.info("=" * 60)

        return True
    else:
        logger.error("Styling process failed - see log for details")
        return False


# ============================================================================
# 脚本入口
# ============================================================================

if __name__ == "__main__":
    import sys

    try:
        success = main()
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("Process interrupted by user")
        sys.exit(1)

    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        sys.exit(1)
