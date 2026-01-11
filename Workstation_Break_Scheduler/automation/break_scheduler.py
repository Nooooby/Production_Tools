"""
休息排班自动化工具

功能:
1. 从 BreakRules 表读取按 DeptID 定义的休息段
2. 支持 BatchCount / BatchSize 两种分批模式
3. 按站位/工种分批安排休息
4. 冲突检查：同一时段休息人数不得超过允许上限
5. 输出休息安排至 Schedule 表
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path
import logging
import math

import openpyxl
from openpyxl.utils import get_column_letter


class Config:
    """系统配置"""

    BASE_DIR = Path(r"C:\Projects\Production_management\Production_Operations_Dashboard")
    INPUT_FILE = BASE_DIR / "data" / "FINAL_v70.xlsx"
    OUTPUT_FILE = BASE_DIR / "data" / "FINAL_v70.xlsx"
    BACKUP_FILE = BASE_DIR / "data" / "FINAL_v70_backup_before_breaks.xlsx"
    LOG_DIR = BASE_DIR / "logs"

    BREAK_RULES_TABLE = "BreakRules"
    SCHEDULE_TABLE = "Schedule"


def setup_logging():
    """配置日志系统"""
    log_dir = Config.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"break_scheduler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )

    return logging.getLogger(__name__)


logger = setup_logging()


@dataclass(frozen=True)
class BreakRule:
    """休息规则"""

    dept_id: str
    break_start: time
    break_end: time
    batch_count: int | None
    batch_size: int | None
    max_concurrent: int | None


@dataclass(frozen=True)
class ScheduleEntry:
    """排班记录"""

    dept_id: str
    employee_id: str
    station: str
    role: str
    start_time: time | None
    end_time: time | None
    entry_type: str


def parse_time(value) -> time | None:
    """解析时间值"""
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        base = datetime(1899, 12, 30) + timedelta(days=value)
        return base.time()
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(value, fmt).time()
            except ValueError:
                continue
    raise ValueError(f"无法解析时间值: {value}")


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def add_minutes(start: time, minutes: float) -> time:
    base = datetime.combine(datetime.today(), start) + timedelta(minutes=minutes)
    return base.time()


def time_minutes(start: time, end: time) -> int:
    start_dt = datetime.combine(datetime.today(), start)
    end_dt = datetime.combine(datetime.today(), end)
    delta = end_dt - start_dt
    return int(delta.total_seconds() // 60)


def overlap(start_a: time, end_a: time, start_b: time, end_b: time) -> bool:
    start_dt_a = datetime.combine(datetime.today(), start_a)
    end_dt_a = datetime.combine(datetime.today(), end_a)
    start_dt_b = datetime.combine(datetime.today(), start_b)
    end_dt_b = datetime.combine(datetime.today(), end_b)
    return start_dt_a < end_dt_b and start_dt_b < end_dt_a


def find_table(workbook, table_name):
    """在工作簿中查找表"""
    for ws in workbook.worksheets:
        if table_name in ws.tables:
            return ws, ws.tables[table_name]
    raise ValueError(f"未找到表: {table_name}")


def read_table_rows(workbook, table_name):
    """读取表格为行字典"""
    ws, table = find_table(workbook, table_name)
    cells = ws[table.ref]
    headers = [normalize_text(cell.value) for cell in cells[0]]
    rows = []
    for row in cells[1:]:
        values = [cell.value for cell in row]
        if all(value in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return ws, table, headers, rows


class BreakScheduler:
    """休息排班器"""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.wb = None

    def load_workbook(self):
        logger.info("加载工作簿: %s", self.workbook_path)
        self.wb = openpyxl.load_workbook(self.workbook_path)

    def create_backup(self, backup_path: Path):
        if backup_path.exists():
            logger.info("备份已存在，跳过: %s", backup_path)
            return
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        backup_path.write_bytes(self.workbook_path.read_bytes())
        logger.info("备份创建成功: %s", backup_path)

    def load_break_rules(self) -> list[BreakRule]:
        _, _, _, rows = read_table_rows(self.wb, Config.BREAK_RULES_TABLE)
        rules = []
        for row in rows:
            dept_id = normalize_text(row.get("DeptID"))
            if not dept_id:
                continue
            break_start = parse_time(row.get("BreakStart"))
            break_end = parse_time(row.get("BreakEnd"))
            batch_count = row.get("BatchCount")
            batch_size = row.get("BatchSize")
            max_concurrent = row.get("MaxConcurrent")
            rules.append(
                BreakRule(
                    dept_id=dept_id,
                    break_start=break_start,
                    break_end=break_end,
                    batch_count=int(batch_count) if batch_count else None,
                    batch_size=int(batch_size) if batch_size else None,
                    max_concurrent=int(max_concurrent) if max_concurrent else None,
                )
            )
        return rules

    def load_schedule_entries(self) -> tuple[list[ScheduleEntry], list[str], object, object]:
        ws, table, headers, rows = read_table_rows(self.wb, Config.SCHEDULE_TABLE)
        entries = []
        for row in rows:
            dept_id = normalize_text(row.get("DeptID"))
            employee_id = normalize_text(row.get("EmployeeID"))
            station = normalize_text(row.get("Station"))
            role = normalize_text(row.get("Role"))
            entry_type = normalize_text(row.get("Type")) or "Work"
            start_time = parse_time(row.get("StartTime") or row.get("Start"))
            end_time = parse_time(row.get("EndTime") or row.get("End"))
            entries.append(
                ScheduleEntry(
                    dept_id=dept_id,
                    employee_id=employee_id,
                    station=station,
                    role=role,
                    start_time=start_time,
                    end_time=end_time,
                    entry_type=entry_type,
                )
            )
        return entries, headers, ws, table

    def resolve_batch_count(self, rule: BreakRule, employee_count: int) -> int:
        if rule.batch_count:
            return max(rule.batch_count, 1)
        if rule.batch_size:
            return max(math.ceil(employee_count / rule.batch_size), 1)
        return 1

    def resolve_batch_size(self, rule: BreakRule, employee_count: int, batch_count: int) -> int:
        if rule.batch_size:
            return max(rule.batch_size, 1)
        return max(math.ceil(employee_count / batch_count), 1)

    def group_employees(self, entries: list[ScheduleEntry]):
        grouped = {}
        for entry in entries:
            if entry.entry_type.lower() == "break":
                continue
            key = (entry.dept_id, entry.station, entry.role)
            grouped.setdefault(key, []).append(entry)
        return grouped

    def build_existing_break_counts(self, entries: list[ScheduleEntry]):
        existing = []
        for entry in entries:
            if entry.entry_type.lower() != "break":
                continue
            if not entry.start_time or not entry.end_time:
                continue
            existing.append(entry)
        return existing

    def insert_breaks(self) -> list[ScheduleEntry]:
        rules = self.load_break_rules()
        entries, _, _, _ = self.load_schedule_entries()
        grouped = self.group_employees(entries)
        existing_breaks = self.build_existing_break_counts(entries)

        new_breaks = []
        for rule in rules:
            if not rule.break_start or not rule.break_end:
                raise ValueError(f"DeptID {rule.dept_id} 缺少 BreakStart/BreakEnd")

            matching_groups = {
                key: value for key, value in grouped.items() if key[0] == rule.dept_id
            }
            if not matching_groups:
                continue

            total_minutes = time_minutes(rule.break_start, rule.break_end)
            if total_minutes <= 0:
                raise ValueError(f"DeptID {rule.dept_id} 休息段时长无效")

            for (_, station, role), employees in matching_groups.items():
                employee_count = len(employees)
                batch_count = self.resolve_batch_count(rule, employee_count)
                batch_size = self.resolve_batch_size(rule, employee_count, batch_count)
                slot_minutes = total_minutes / batch_count
                max_concurrent = rule.max_concurrent or batch_size

                sorted_employees = sorted(
                    employees, key=lambda item: (item.station, item.role, item.employee_id)
                )

                for batch_index in range(batch_count):
                    batch_start = add_minutes(rule.break_start, slot_minutes * batch_index)
                    batch_end = add_minutes(
                        rule.break_start, slot_minutes * (batch_index + 1)
                    )
                    batch_members = sorted_employees[
                        batch_index * batch_size : (batch_index + 1) * batch_size
                    ]
                    if not batch_members:
                        continue

                    existing_count = sum(
                        1
                        for entry in existing_breaks
                        if entry.dept_id == rule.dept_id
                        and entry.start_time
                        and entry.end_time
                        and overlap(entry.start_time, entry.end_time, batch_start, batch_end)
                    )
                    proposed_count = existing_count + len(batch_members)
                    if proposed_count > max_concurrent:
                        raise ValueError(
                            f"DeptID {rule.dept_id} 在 {batch_start}-{batch_end} 休息人数超限"
                        )

                    for member in batch_members:
                        new_breaks.append(
                            ScheduleEntry(
                                dept_id=member.dept_id,
                                employee_id=member.employee_id,
                                station=station,
                                role=role,
                                start_time=batch_start,
                                end_time=batch_end,
                                entry_type="Break",
                            )
                        )

        return new_breaks

    def write_breaks(self, breaks: list[ScheduleEntry]):
        if not breaks:
            logger.info("未生成休息安排")
            return

        _, headers, ws, table = self.load_schedule_entries()
        header_map = {header: idx for idx, header in enumerate(headers)}

        start_col = min(cell.col_idx for cell in ws[table.ref][0])
        end_col = start_col + len(headers) - 1

        start_row = ws[table.ref][0][0].row
        end_row = ws[table.ref][-1][0].row

        def build_row(entry: ScheduleEntry):
            values = [""] * len(headers)
            for name, value in [
                ("DeptID", entry.dept_id),
                ("EmployeeID", entry.employee_id),
                ("Station", entry.station),
                ("Role", entry.role),
                ("Type", entry.entry_type),
                ("StartTime", entry.start_time),
                ("EndTime", entry.end_time),
                ("Start", entry.start_time),
                ("End", entry.end_time),
            ]:
                if name in header_map and value is not None:
                    values[header_map[name]] = value
            return values

        for entry in breaks:
            end_row += 1
            row_values = build_row(entry)
            for offset, value in enumerate(row_values):
                ws.cell(row=end_row, column=start_col + offset, value=value)

        table.ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        logger.info("写入休息安排: %s 条", len(breaks))

    def run(self):
        self.load_workbook()
        self.create_backup(Config.BACKUP_FILE)
        breaks = self.insert_breaks()
        self.write_breaks(breaks)
        self.wb.save(self.workbook_path)
        logger.info("休息排班完成")


def InsertBreaks(workbook_path: Path | None = None):
    """入口函数"""
    path = workbook_path or Config.INPUT_FILE
    scheduler = BreakScheduler(Path(path))
    scheduler.run()


if __name__ == "__main__":
    InsertBreaks()
