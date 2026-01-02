# -*- coding: utf-8 -*-
"""
fix_cages_max_logic.py - 修正 Cages 汇总逻辑为取最大值

核心修改：
1. 在 14_Production_Planning 添加按部位分组区域（Row 20-35）
2. 修改 G9 从 SUM 改为引用 MAX 结果

原理：
- 切割一个笼子会同时产出所有部位（Breast, ThighMeat, Drum, Wing 等）
- 所以总 Cages = MAX(所有部位的 Cages)，而不是 SUM

作者: Claude
日期: 2026-01-02
"""

import openpyxl
from datetime import datetime
import os

# 配置
INPUT_FILE = 'Production_Operations_Dashboard/data/v39_Dashboard_Enhanced.xlsx'
OUTPUT_FILE = 'Production_Operations_Dashboard/data/v39_Dashboard_Enhanced.xlsx'
LOG_FILE = f'Production_Operations_Dashboard/logs/fix_cages_max_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

os.makedirs('Production_Operations_Dashboard/logs', exist_ok=True)

def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    with open(LOG_FILE, 'a', encoding='utf-8')  as f:
        f.write(log_line + '\n')

def get_product_groups(wb):
    """从 00_Yield_Rates 获取所有部位列表"""
    ws = wb['00_Yield_Rates']
    product_groups = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value  # B: Short Name
        if name and name not in product_groups:
            product_groups.append(name)

    log(f"Found {len(product_groups)} product groups: {', '.join(product_groups)}")
    return product_groups

def add_by_part_aggregation(wb, product_groups):
    """
    在 14_Production_Planning 添加按部位分组汇总区域

    Row 20: 标题
    Row 21: 列头
    Row 22-34: 每个部位一行
    Row 35: 总 Cages（MAX）
    """
    ws = wb['14_Production_Planning']

    # Row 20: 标题
    ws['A20'] = '四、按部位汇总 Cages 需求'
    ws['A20'].font = openpyxl.styles.Font(bold=True, size=12)

    # Row 21: 列头
    headers = ['部位 (Product_Group)', 'TrayPack Cages', 'BulkPack Cages', 'Bagging Cages', '总 Cages']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(21, col)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    log(f"Added headers at Row 21")

    # Row 22-34: 每个部位
    row_num = 22
    for pg in product_groups:
        # A: Product_Group 名称
        ws.cell(row_num, 1).value = pg

        # B: TrayPack Cages (SUMIFS from 05_Daily_Orders)
        ws.cell(row_num, 2).value = f'=SUMIFS(\'05_Daily_Orders\'!$S$2:$S$328,\'05_Daily_Orders\'!$N$2:$N$328,\"{pg}\")'

        # C: BulkPack Cages (SUMIFS from 10_Cone_Line)
        ws.cell(row_num, 3).value = f'=SUMIFS(\'10_Cone_Line\'!$N$2:$N$129,\'10_Cone_Line\'!$I$2:$I$129,\"{pg}\")'

        # D: Bagging Cages (SUMIFS from 04_Bagging_Order)
        ws.cell(row_num, 4).value = f'=SUMIFS(\'04_Bagging_Order\'!$S$5:$S$22,\'04_Bagging_Order\'!$N$5:$N$22,\"{pg}\")'

        # E: Total = B + C + D
        ws.cell(row_num, 5).value = f'=B{row_num}+C{row_num}+D{row_num}'

        row_num += 1

    log(f"Added {len(product_groups)} product group rows (Row 22-{row_num-1})")

    # Row 35 (或最后一行+1): 总 Cages 需求 = MAX
    max_row = row_num
    ws.cell(max_row, 1).value = '总 Cages 需求（MAX）'
    ws.cell(max_row, 1).font = openpyxl.styles.Font(bold=True)
    ws.cell(max_row, 5).value = f'=MAX(E22:E{row_num-1})'
    ws.cell(max_row, 5).font = openpyxl.styles.Font(bold=True)
    ws.cell(max_row, 5).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    log(f"Added MAX formula at Row {max_row}: E{max_row}")

    return max_row

def update_g9_formula(wb, max_result_cell):
    """修改 G9 从 SUM 改为引用 MAX 结果"""
    ws = wb['14_Production_Planning']

    old_formula = ws['G9'].value
    log(f"Old G9 formula: {old_formula}")

    # 修改 G9 引用新的 MAX 结果
    ws['G9'].value = f'=E{max_result_cell}'

    log(f"Updated G9 formula: =E{max_result_cell}")

def main():
    log("=" * 60)
    log("Fix Cages MAX Logic - Start")
    log("=" * 60)

    # 加载工作簿
    log(f"Loading workbook: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # 1. 获取所有部位列表
    product_groups = get_product_groups(wb)

    # 2. 添加按部位分组汇总区域
    max_result_row = add_by_part_aggregation(wb, product_groups)

    # 3. 更新 G9 公式
    update_g9_formula(wb, max_result_row)

    # 保存
    log(f"Saving workbook: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    wb.close()

    log("=" * 60)
    log("Fix Cages MAX Logic - Complete")
    log(f"Log file: {LOG_FILE}")
    log("=" * 60)

    return True

if __name__ == '__main__':
    success = main()
    exit(0 if success else 1)
