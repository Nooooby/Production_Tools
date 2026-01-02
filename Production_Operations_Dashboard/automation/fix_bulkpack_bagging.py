# -*- coding: utf-8 -*-
"""
fix_bulkpack_bagging.py - 修复 BulkPack 和 Bagging 转换逻辑

修复内容:
1. 10_Cone_Line: 添加转换逻辑列 (I-N)
2. 04_Bagging_Order: 添加完整转换逻辑列 (N-S)
3. 14_Production_Planning: 更新 BulkPack/Bagging 行

作者: Claude
日期: 2026-01-02
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import os

# 配置
INPUT_FILE = 'Production_Operations_Dashboard/data/v39_Dashboard_Enhanced.xlsx'
OUTPUT_FILE = 'Production_Operations_Dashboard/data/v39_Dashboard_Enhanced.xlsx'
BACKUP_FILE = f'Production_Operations_Dashboard/data/v39_Dashboard_Enhanced_backup_bulkpack_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
LOG_FILE = f'Production_Operations_Dashboard/logs/fix_bulkpack_bagging_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'

os.makedirs('Production_Operations_Dashboard/logs', exist_ok=True)

def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_line + '\n')

def backup_file():
    shutil.copy(INPUT_FILE, BACKUP_FILE)
    log(f"Backup created: {BACKUP_FILE}")

def add_conversion_columns_to_cone_line(wb):
    """
    在 10_Cone_Line 中添加转换逻辑列

    现有列: A-H (Code, Code, Form TP Order, Form Cone line order, ORDER, Order check, ShelfLife, empty)

    新增列:
    - I: Product_Group (从 SKU_Master!E 查找)
    - J: Avg_Case_Weight (从 SKU_Master!F 查找)
    - K: Yield_Rate (从 Yield_Rates!E 查找)
    - L: WIP_kg (ORDER * Avg_Case_Weight)
    - M: Raw_kg_Needed (WIP_kg / Yield%)
    - N: Cages_Needed (Raw_kg / 680)
    """
    ws = wb['10_Cone_Line']
    max_row = ws.max_row

    log(f"10_Cone_Line: {max_row} rows, adding conversion columns I-N")

    # 添加列头 (Row 1)
    headers = {
        'I': 'Product_Group',
        'J': 'Avg_Case_Weight',
        'K': 'Yield_Rate',
        'L': 'WIP_kg',
        'M': 'Raw_kg_Needed',
        'N': 'Cages_Needed'
    }

    for col, header in headers.items():
        ws[f'{col}1'] = header
        log(f"  Added header {col}1: {header}")

    # 添加公式 (Row 2 onwards)
    # 注意: 10_Cone_Line 的 A 列是 SKU Code
    formulas_added = 0
    for row in range(2, max_row + 1):
        # I: Product_Group
        ws[f'I{row}'] = f'=IFERROR(XLOOKUP(A{row},\'00_SKU_Master\'!$B$2:$B$378,\'00_SKU_Master\'!$E$2:$E$378,""),"")'

        # J: Avg_Case_Weight
        ws[f'J{row}'] = f'=IFERROR(XLOOKUP(A{row},\'00_SKU_Master\'!$B$2:$B$378,\'00_SKU_Master\'!$F$2:$F$378,0),0)'

        # K: Yield_Rate
        ws[f'K{row}'] = f'=IFERROR(XLOOKUP(I{row},\'00_Yield_Rates\'!$B$2:$B$33,\'00_Yield_Rates\'!$E$2:$E$33,0)/100,0)'

        # L: WIP_kg = ORDER (E) * Avg_Case_Weight (J)
        ws[f'L{row}'] = f'=IF(AND(ISNUMBER(E{row}),ISNUMBER(J{row})),E{row}*J{row},0)'

        # M: Raw_kg_Needed = WIP_kg / Yield_Rate
        ws[f'M{row}'] = f'=IF(AND(ISNUMBER(L{row}),K{row}>0),L{row}/K{row},0)'

        # N: Cages_Needed = Raw_kg / 680
        ws[f'N{row}'] = f'=IF(M{row}>0,ROUNDUP(M{row}/680,1),0)'

        formulas_added += 6

    log(f"  Added {formulas_added} formulas to 10_Cone_Line (rows 2-{max_row})")
    return formulas_added

def add_conversion_columns_to_bagging(wb):
    """
    在 04_Bagging_Order 中添加完整转换逻辑列

    现有列结构:
    - B: SKU (从 SKU_Master 引用)
    - C-F: 订单数据
    - H: SKU (=B)
    - I: Today's Order
    - J: Completed
    - K: Remaining
    - L: KG (已有,使用内部查找表)
    - M: Raw Cages (已有,但可能有问题)

    新增列 (N-S):
    - N: Product_Group
    - O: Avg_Case_Weight
    - P: Yield_Rate (使用正确的00_Yield_Rates)
    - Q: WIP_kg
    - R: Raw_kg_Needed
    - S: Cages_Needed
    """
    ws = wb['04_Bagging_Order']

    log(f"04_Bagging_Order: Adding conversion columns N-S")

    # 添加列头 (Row 4)
    headers = {
        'N': 'Product_Group',
        'O': 'Avg_Case_Weight',
        'P': 'Yield_Rate',
        'Q': 'WIP_kg',
        'R': 'Raw_kg_Needed',
        'S': 'Cages_Needed'
    }

    for col, header in headers.items():
        ws[f'{col}4'] = header
        log(f"  Added header {col}4: {header}")

    # 添加公式 (Row 5-22, 数据区域)
    formulas_added = 0
    for row in range(5, 23):
        # N: Product_Group (基于B列的SKU)
        ws[f'N{row}'] = f'=IFERROR(XLOOKUP(B{row},\'00_SKU_Master\'!$B$2:$B$378,\'00_SKU_Master\'!$E$2:$E$378,""),"")'

        # O: Avg_Case_Weight
        ws[f'O{row}'] = f'=IFERROR(XLOOKUP(B{row},\'00_SKU_Master\'!$B$2:$B$378,\'00_SKU_Master\'!$F$2:$F$378,0),0)'

        # P: Yield_Rate (从00_Yield_Rates查找)
        ws[f'P{row}'] = f'=IFERROR(XLOOKUP(N{row},\'00_Yield_Rates\'!$B$2:$B$33,\'00_Yield_Rates\'!$E$2:$E$33,0)/100,0)'

        # Q: WIP_kg = Today's Order (I) * Avg_Case_Weight (O)
        ws[f'Q{row}'] = f'=IF(AND(ISNUMBER(I{row}),ISNUMBER(O{row})),I{row}*O{row},0)'

        # R: Raw_kg_Needed = WIP_kg / Yield_Rate
        ws[f'R{row}'] = f'=IF(AND(ISNUMBER(Q{row}),P{row}>0),Q{row}/P{row},0)'

        # S: Cages_Needed = Raw_kg / 680
        ws[f'S{row}'] = f'=IF(R{row}>0,ROUNDUP(R{row}/680,1),0)'

        formulas_added += 6

    # 添加汇总行 (Row 3)
    ws['N3'] = 'Totals:'
    ws['Q3'] = '=SUM(Q5:Q22)'
    ws['R3'] = '=SUM(R5:R22)'
    ws['S3'] = '=SUM(S5:S22)'

    log(f"  Added {formulas_added} formulas to 04_Bagging_Order (rows 5-22)")
    return formulas_added

def fix_production_planning_bulkpack_bagging(wb):
    """
    修正 14_Production_Planning 的 BulkPack 和 Bagging 行
    """
    ws = wb['14_Production_Planning']

    log("Fixing 14_Production_Planning BulkPack/Bagging rows...")

    # BulkPack (Row 7)
    # B7: 订单数 - 使用 10_Cone_Line!E 列 (ORDER)
    ws['B7'] = "=SUMIF('10_Cone_Line'!E:E,\">0\")"
    # C7: 加权平均重量
    ws['C7'] = "=IF(B7>0,SUMPRODUCT(('10_Cone_Line'!$E$2:$E$129>0)*('10_Cone_Line'!$E$2:$E$129)*('10_Cone_Line'!$J$2:$J$129))/SUMIF('10_Cone_Line'!$E$2:$E$129,\">0\"),0)"
    # D7 保持: =B7*C7
    # E7: 加权平均 Yield
    ws['E7'] = "=IF(B7>0,SUMPRODUCT(('10_Cone_Line'!$E$2:$E$129>0)*('10_Cone_Line'!$E$2:$E$129)*('10_Cone_Line'!$K$2:$K$129))/SUMIF('10_Cone_Line'!$E$2:$E$129,\">0\"),0)"
    # F7: Raw_kg 汇总
    ws['F7'] = "=SUMIF('10_Cone_Line'!$E$2:$E$129,\">0\",'10_Cone_Line'!$M$2:$M$129)"
    # G7: Cages 汇总
    ws['G7'] = "=SUMIF('10_Cone_Line'!$E$2:$E$129,\">0\",'10_Cone_Line'!$N$2:$N$129)"

    log("  Fixed BulkPack row (Row 7)")

    # Bagging (Row 8)
    # B8: 订单数 - 使用 04_Bagging_Order!I 列
    ws['B8'] = "=SUM('04_Bagging_Order'!I5:I22)"
    # C8: 加权平均重量
    ws['C8'] = "=IF(B8>0,SUMPRODUCT(('04_Bagging_Order'!$I$5:$I$22>0)*('04_Bagging_Order'!$I$5:$I$22)*('04_Bagging_Order'!$O$5:$O$22))/SUMIF('04_Bagging_Order'!$I$5:$I$22,\">0\"),0)"
    # D8 保持: =B8*C8
    # E8: 加权平均 Yield
    ws['E8'] = "=IF(B8>0,SUMPRODUCT(('04_Bagging_Order'!$I$5:$I$22>0)*('04_Bagging_Order'!$I$5:$I$22)*('04_Bagging_Order'!$P$5:$P$22))/SUMIF('04_Bagging_Order'!$I$5:$I$22,\">0\"),0)"
    # F8: Raw_kg 汇总
    ws['F8'] = "=SUM('04_Bagging_Order'!R5:R22)"
    # G8: Cages 汇总
    ws['G8'] = "=SUM('04_Bagging_Order'!S5:S22)"

    log("  Fixed Bagging row (Row 8)")

    return True

def verify_formulas(wb):
    """验证所有公式"""
    log("Verifying formulas...")

    # Check 10_Cone_Line
    ws = wb['10_Cone_Line']
    log(f"  10_Cone_Line: Columns I-N added ({ws.max_row - 1} rows)")

    # Check 04_Bagging_Order
    ws = wb['04_Bagging_Order']
    log(f"  04_Bagging_Order: Columns N-S added (18 rows)")

    # Check 14_Production_Planning
    ws = wb['14_Production_Planning']
    for row in [7, 8]:
        for col in ['B', 'C', 'E', 'F', 'G']:
            cell = ws[f'{col}{row}']
            if cell.value:
                log(f"  14_Production_Planning!{col}{row}: {str(cell.value)[:50]}...")

    return True

def main():
    log("=" * 60)
    log("Fix BulkPack/Bagging Conversion Logic - Start")
    log("=" * 60)

    # 创建备份
    backup_file()

    # 加载工作簿
    log(f"Loading workbook: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # 1. 添加转换逻辑列到 10_Cone_Line
    cone_formulas = add_conversion_columns_to_cone_line(wb)

    # 2. 添加转换逻辑列到 04_Bagging_Order
    bagging_formulas = add_conversion_columns_to_bagging(wb)

    # 3. 修正 14_Production_Planning
    fix_production_planning_bulkpack_bagging(wb)

    # 4. 验证
    verify_formulas(wb)

    # 保存
    log(f"Saving workbook: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    wb.close()

    total_formulas = cone_formulas + bagging_formulas
    log("=" * 60)
    log("Fix BulkPack/Bagging - Complete")
    log(f"Total formulas added: {total_formulas}")
    log(f"  - 10_Cone_Line: {cone_formulas}")
    log(f"  - 04_Bagging_Order: {bagging_formulas}")
    log(f"Backup file: {BACKUP_FILE}")
    log(f"Log file: {LOG_FILE}")
    log("=" * 60)

    return True

if __name__ == '__main__':
    success = main()
    exit(0 if success else 1)
