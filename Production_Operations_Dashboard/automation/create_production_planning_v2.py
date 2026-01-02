"""
14_Production_Planning v2 - 完整生产物料规划表
Production Material Planning Sheet with Complete Cages Calculation

功能:
1. 聚合三类订单 (TrayPack + BulkPack + Bagging)
2. 计算每类订单需要多少 Cages
3. 显示库存和鸟笼需求对比
4. 计算原料库存状态

数据源:
- 05_Daily_Orders Column M (TrayPack 订单 cases)
- 10_Cone_Line Column M (BulkPack 订单 cases)
- 04_Bagging_Order I5:I22 (Bagging 订单 cases)
- 00_SKU_Master (E=Product_Group, F=Avg_Case_Weight)
- 00_Yield_Rates (B=Product, E=Yield%)
- 01_Cages_Plan C3 (可切笼数)
- 06_Resource_Plan (原料库存)

转换公式:
WIP kg = Cases × Avg_Case_Weight
Raw kg needed = WIP kg ÷ Yield%
Cages needed = Raw kg needed ÷ 680 kg/cage

作者: Claude Code
创建日期: 2026-01-01
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from datetime import datetime
import shutil

# ============================================================================
# 配置
# ============================================================================

class Config:
    """系统配置"""

    BASE_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard')
    INPUT_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced.xlsx'
    OUTPUT_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced.xlsx'
    BACKUP_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced_backup_before_planning_v2.xlsx'
    LOG_DIR = BASE_DIR / 'logs'

    # 颜色定义
    COLOR_HEADER = '343A40'
    COLOR_GOOD = '70AD47'
    COLOR_WARNING = 'FFC000'
    COLOR_CRITICAL = 'FF6B6B'
    COLOR_BG_LIGHT = 'F8F9FA'

    # 生产参数
    KG_PER_CAGE = 680  # 平均每笼的肉类产出

# ============================================================================
# 日志设置
# ============================================================================

def setup_logging():
    """配置日志系统"""
    log_dir = Config.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"production_planning_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

logger = setup_logging()

# ============================================================================
# 生产规划表生成器 v2
# ============================================================================

class ProductionPlanningGeneratorV2:
    """生成和更新 14_Production_Planning 工作表 - 完整版"""

    def __init__(self, workbook_path):
        """初始化"""
        self.workbook_path = workbook_path
        self.wb = None
        self.ws = None

        logger.info("初始化生产规划表生成器 v2")

    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"加载工作簿: {self.workbook_path}")
        self.wb = openpyxl.load_workbook(self.workbook_path)

        # 获取或创建工作表
        if '14_Production_Planning' in self.wb.sheetnames:
            self.ws = self.wb['14_Production_Planning']
            # 清空现有内容（包括处理合并单元格）
            for merged_range in list(self.ws.merged_cells.ranges):
                self.ws.unmerge_cells(str(merged_range))
            for row in self.ws.iter_rows():
                for cell in row:
                    cell.value = None
            logger.info("更新现有的 14_Production_Planning 工作表")
        else:
            self.ws = self.wb.create_sheet('14_Production_Planning', index=13)
            logger.info("创建新的 14_Production_Planning 工作表")

    def create_backup(self, backup_path):
        """创建备份"""
        logger.info(f"创建备份: {backup_path}")
        shutil.copy2(self.workbook_path, backup_path)
        logger.info("备份创建成功")

    def add_title(self):
        """添加标题"""
        self.ws['A1'] = "生产物料规划表 - 订单与鸡笼需求分析"
        self.ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        self.ws['A1'].fill = PatternFill(start_color=Config.COLOR_HEADER,
                                        end_color=Config.COLOR_HEADER, fill_type='solid')
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 25
        self.ws.merge_cells('A1:H1')

        # 日期
        self.ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        self.ws['A2'].font = Font(name='Calibri', size=10, italic=True)

        logger.info("添加标题完成")

    def add_order_and_cages_summary(self):
        """添加订单与鸡笼需求汇总"""
        logger.info("添加订单与鸡笼需求汇总...")

        start_row = 4

        # 标题
        self.ws[f'A{start_row}'] = "一、订单与鸡笼需求"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['订单类型', '订单数(cases)', 'Avg重量(kg)', 'WIP需求(kg)',
                   'Yield(%)', '原料需求(kg)', 'Cages需要', '状态']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row = start_row + 2

        # TrayPack
        self.ws[f'A{row}'] = "TrayPack"
        self.ws[f'B{row}'] = "=SUMIF('05_Daily_Orders'!M:M,\">0\")"
        self.ws[f'C{row}'] = "=AVERAGE('00_SKU_Master'!F:F)"
        self.ws[f'D{row}'] = f"=B{row}*C{row}"
        self.ws[f'E{row}'] = "=AVERAGE('00_Yield_Rates'!E:E)"
        self.ws[f'F{row}'] = f"=IF(E{row}=0,0,D{row}/E{row}*100)"
        self.ws[f'G{row}'] = f"=IF(F{row}=0,0,F{row}/{Config.KG_PER_CAGE})"
        row += 1

        # BulkPack
        self.ws[f'A{row}'] = "BulkPack"
        self.ws[f'B{row}'] = "=SUMIF('10_Cone_Line'!M:M,\">0\")"
        self.ws[f'C{row}'] = "=AVERAGE('00_SKU_Master'!F:F)"
        self.ws[f'D{row}'] = f"=B{row}*C{row}"
        self.ws[f'E{row}'] = "=AVERAGE('00_Yield_Rates'!E:E)"
        self.ws[f'F{row}'] = f"=IF(E{row}=0,0,D{row}/E{row}*100)"
        self.ws[f'G{row}'] = f"=IF(F{row}=0,0,F{row}/{Config.KG_PER_CAGE})"
        row += 1

        # Bagging
        self.ws[f'A{row}'] = "Bagging"
        self.ws[f'B{row}'] = "=SUM('04_Bagging_Order'!I5:I22)"
        self.ws[f'C{row}'] = "=AVERAGE('00_SKU_Master'!F:F)"
        self.ws[f'D{row}'] = f"=B{row}*C{row}"
        self.ws[f'E{row}'] = "=AVERAGE('00_Yield_Rates'!E:E)"
        self.ws[f'F{row}'] = f"=IF(E{row}=0,0,D{row}/E{row}*100)"
        self.ws[f'G{row}'] = f"=IF(F{row}=0,0,F{row}/{Config.KG_PER_CAGE})"
        row += 1

        # 合计
        self.ws[f'A{row}'] = "总需要 Cages"
        self.ws[f'A{row}'].font = Font(bold=True, size=11)
        self.ws[f'G{row}'] = f"=SUM(G{row-3}:G{row-1})"
        self.ws[f'G{row}'].font = Font(bold=True, size=11)
        self.ws[f'G{row}'].fill = PatternFill(start_color=Config.COLOR_GOOD,
                                             end_color=Config.COLOR_GOOD, fill_type='solid')
        row += 2

        logger.info("订单与鸡笼需求汇总添加完成")
        return row

    def add_cage_availability(self, start_row):
        """添加鸡笼库存与需求对比"""
        logger.info("添加鸡笼库存与需求对比...")

        # 标题
        self.ws[f'A{start_row}'] = "二、鸡笼库存与需求对比"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['项目', '数量', '说明']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')

        row = start_row + 2

        # 可用笼数
        self.ws[f'A{row}'] = "今日可切笼数"
        self.ws[f'B{row}'] = "='01_Cages_Plan'!C3"
        self.ws[f'C{row}'] = "从鸡笼计划"
        available_row = row
        row += 1

        # 需要笼数
        self.ws[f'A{row}'] = "所有订单需要笼数"
        prev_total_cages_row = start_row - 2
        self.ws[f'B{row}'] = f"=G{prev_total_cages_row}"
        self.ws[f'C{row}'] = "TrayPack + BulkPack + Bagging"
        needed_row = row
        row += 1

        # 剩余笼数
        self.ws[f'A{row}'] = "剩余笼数"
        self.ws[f'B{row}'] = f"=B{available_row}-B{needed_row}"
        self.ws[f'C{row}'] = ""
        remaining_row = row
        row += 1

        # 库存状态
        self.ws[f'A{row}'] = "库存状态"
        self.ws[f'B{row}'] = f"=IF(B{remaining_row}>=0,\"充足\",\"不足\")"
        self.ws[f'C{row}'] = "缺口数量"

        logger.info("鸡笼库存与需求对比添加完成")
        return row + 2

    def add_raw_material_summary(self, start_row):
        """添加原料库存汇总"""
        logger.info("添加原料库存汇总...")

        # 标题
        self.ws[f'A{start_row}'] = "三、原料库存状态"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['原料SKU', '库存(cases)', '需求(cases)', '剩余', '状态']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')

        row = start_row + 2

        # 从 06_Resource_Plan 读取原料数据
        for i in range(2, 43):
            self.ws[f'A{row}'] = f"='06_Resource_Plan'!C{i}"
            self.ws[f'B{row}'] = f"='06_Resource_Plan'!L{i}"
            self.ws[f'C{row}'] = f"='06_Resource_Plan'!D{i}"
            self.ws[f'D{row}'] = f"=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),B{row}-C{row},\"\")"
            self.ws[f'E{row}'] = f"=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),IF(B{row}>=C{row},\"OK\",\"缺货\"),\"\")"

            row += 1

        logger.info("原料库存汇总添加完成")
        return row + 2

    def optimize_columns(self):
        """优化列宽"""
        logger.info("优化列宽...")

        widths = {
            'A': 20, 'B': 15, 'C': 15, 'D': 15, 'E': 12, 'F': 15, 'G': 15, 'H': 15
        }
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

    def apply_formatting(self):
        """应用格式"""
        logger.info("应用格式...")

        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row,
                                     min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def save_workbook(self):
        """保存工作簿"""
        logger.info(f"保存工作簿到: {Config.OUTPUT_FILE}")
        self.wb.save(str(Config.OUTPUT_FILE))
        logger.info("工作簿保存成功")

    def run(self):
        """执行完整流程"""
        try:
            logger.info("=" * 80)
            logger.info("开始生成生产规划表 v2（完整版）")
            logger.info("=" * 80)

            # 加载工作簿
            self.load_workbook()

            # 创建备份
            self.create_backup(Config.BACKUP_FILE)

            # 生成内容
            self.add_title()
            next_row = self.add_order_and_cages_summary()
            next_row = self.add_cage_availability(next_row)
            next_row = self.add_raw_material_summary(next_row)

            # 格式化
            self.optimize_columns()
            self.apply_formatting()

            # 保存
            self.save_workbook()

            logger.info("=" * 80)
            logger.info("生产规划表 v2 生成成功")
            logger.info("=" * 80)
            logger.info(f"输出文件: {Config.OUTPUT_FILE}")
            logger.info(f"备份文件: {Config.BACKUP_FILE}")

            return True

        except Exception as e:
            logger.error(f"生成过程出错: {str(e)}", exc_info=True)
            return False

# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数"""
    generator = ProductionPlanningGeneratorV2(Config.INPUT_FILE)
    success = generator.run()

    if success:
        print(f"\nProduction Planning Table v2 generated successfully")
        print(f"File: {Config.OUTPUT_FILE}")
    else:
        print(f"\nGeneration failed, check logs")

    return 0 if success else 1

if __name__ == "__main__":
    exit(main())
