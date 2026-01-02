"""
Dashboard 快速增强系统 - Phase 3 实现
Executive Dashboard Enhancement System

功能:
1. 添加 6 个关键 KPI 指标
2. 创建 3 个可视化图表
3. 应用条件格式规则
4. 自动颜色编码状态显示

作者: Claude Code
创建日期: 2026-01-01
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, DoughnutChart
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from pathlib import Path
import logging
from datetime import datetime
import shutil


# ============================================================================
# 配置和常量
# ============================================================================

class Config:
    """系统配置"""

    # 文件路径
    BASE_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard')
    INPUT_FILE = BASE_DIR / 'data' / 'v39_Normalized_Colored.xlsx'
    OUTPUT_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced.xlsx'
    BACKUP_FILE = BASE_DIR / 'data' / 'v39_Normalized_Colored_backup_before_dashboard.xlsx'
    LOG_DIR = BASE_DIR / 'logs'

    # 工作表名称
    DASHBOARD_SHEET = '12_Executive_Dash'

    # 颜色定义
    KPI_HEADER_BG = '343A40'      # 深灰
    KPI_HEADER_FG = 'FFFFFF'      # 白色
    STATUS_GREEN = '70AD47'       # 绿色
    STATUS_YELLOW = 'FFC000'      # 黄色
    STATUS_RED = 'FF6B6B'         # 红色
    STATUS_LIGHT_GRAY = 'D9D9D9'  # 浅灰

    # KPI 位置
    KPI_START_COL = 15  # 列 O


# ============================================================================
# 日志设置
# ============================================================================

def setup_logging():
    """配置日志系统"""
    log_dir = Config.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"dashboard_enhancement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

logger = setup_logging()


# ============================================================================
# 样式工厂类
# ============================================================================

class StyleFactory:
    """创建样式对象"""

    @staticmethod
    def create_kpi_label_font():
        """KPI 标签字体"""
        return Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    @staticmethod
    def create_kpi_value_font():
        """KPI 数值字体"""
        return Font(name='Calibri', size=18, bold=True, color='000000')

    @staticmethod
    def create_kpi_label_fill():
        """KPI 标签填充"""
        return PatternFill(start_color=Config.KPI_HEADER_BG,
                          end_color=Config.KPI_HEADER_BG,
                          fill_type='solid')

    @staticmethod
    def create_kpi_value_fill(color):
        """KPI 数值填充"""
        return PatternFill(start_color=color,
                          end_color=color,
                          fill_type='solid')

    @staticmethod
    def create_center_alignment():
        """居中对齐"""
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def create_border():
        """边框"""
        side = Side(style='thin', color='CED4DA')
        return Border(left=side, right=side, top=side, bottom=side)


# ============================================================================
# 主 Dashboard 增强类
# ============================================================================

class DashboardEnhancer:
    """Dashboard 增强器"""

    def __init__(self, workbook_path):
        """初始化"""
        self.workbook_path = workbook_path
        self.wb = None
        self.ws = None

    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"加载工作簿: {self.workbook_path}")
        self.wb = openpyxl.load_workbook(self.workbook_path)
        self.ws = self.wb[Config.DASHBOARD_SHEET]
        logger.info(f"工作表加载完成: {Config.DASHBOARD_SHEET}")

    def create_backup(self, backup_path):
        """创建备份"""
        logger.info(f"创建备份: {backup_path}")
        shutil.copy2(self.workbook_path, backup_path)
        logger.info("备份创建成功")

    def add_kpi_section(self):
        """添加 KPI 指标区域"""
        logger.info("开始添加 KPI 指标...")

        col = Config.KPI_START_COL

        # 标题
        title_cell = self.ws.cell(row=2, column=col)
        title_cell.value = "关键指标"
        title_cell.font = StyleFactory.create_kpi_label_font()
        title_cell.fill = StyleFactory.create_kpi_label_fill()
        title_cell.alignment = StyleFactory.create_center_alignment()
        title_cell.border = StyleFactory.create_border()

        # 调整列宽
        self.ws.column_dimensions['O'].width = 20
        self.ws.column_dimensions['P'].width = 15

        # KPI 数据
        kpis = [
            ("总完成率", "=(F8+C9+I8)/(F7+C8+I7)*100", 3),
            ("员工到岗率", "=SUM(M3:M8)/SUM(L3:L8)*100", 5),
            ("Tray Pack", "=F8/F7*100", 7),
            ("Cut-Up", "=C9/C8*100", 9),
            ("Bagging", "=I8/I7*100", 11),
            ("员工缺口", "=SUM(L3:L8)-SUM(M3:M8)", 13),
        ]

        for label, formula, row in kpis:
            # 标签
            label_cell = self.ws.cell(row=row, column=col)
            label_cell.value = label
            label_cell.font = StyleFactory.create_kpi_label_font()
            label_cell.fill = StyleFactory.create_kpi_label_fill()
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            label_cell.border = StyleFactory.create_border()

            # 数值
            value_cell = self.ws.cell(row=row, column=col + 1)
            value_cell.value = formula
            value_cell.font = StyleFactory.create_kpi_value_font()
            value_cell.fill = PatternFill(start_color='FFFFFF',
                                         end_color='FFFFFF',
                                         fill_type='solid')
            value_cell.alignment = StyleFactory.create_center_alignment()
            value_cell.border = StyleFactory.create_border()

        logger.info(f"KPI 指标添加完成: 6 个指标")

    def create_production_chart(self):
        """创建生产完成度图表"""
        logger.info("创建生产完成度图表...")

        # 创建辅助数据
        chart_data_start_row = 18
        chart_data_start_col = 2

        # 部门数据
        departments = ['Tray Pack', 'Cut-Up', 'Bagging']
        completed = [8, 9, 8]  # 列号
        targets = [7, 8, 7]    # 列号

        self.ws.cell(row=chart_data_start_row, column=chart_data_start_col).value = "部门"
        self.ws.cell(row=chart_data_start_row, column=chart_data_start_col + 1).value = "已完成"
        self.ws.cell(row=chart_data_start_row, column=chart_data_start_col + 2).value = "目标"

        for i, (dept, comp_col, target_col) in enumerate(zip(departments, completed, targets)):
            row = chart_data_start_row + i + 1
            self.ws.cell(row=row, column=chart_data_start_col).value = dept
            # 公式引用：从相应列获取已完成和目标值
            if comp_col == 8:  # Tray Pack completed
                self.ws.cell(row=row, column=chart_data_start_col + 1).value = f"=F{8}"
                self.ws.cell(row=row, column=chart_data_start_col + 2).value = f"=F{7}"
            elif comp_col == 9:  # Cut-Up completed
                self.ws.cell(row=row, column=chart_data_start_col + 1).value = f"=C{9}"
                self.ws.cell(row=row, column=chart_data_start_col + 2).value = f"=C{8}"
            else:  # Bagging
                self.ws.cell(row=row, column=chart_data_start_col + 1).value = f"=I{8}"
                self.ws.cell(row=row, column=chart_data_start_col + 2).value = f"=I{7}"

        # 创建图表
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "生产完成度"
        chart.y_axis.title = "数量"
        chart.x_axis.title = "部门"

        # 数据范围
        data = Reference(self.ws, min_col=chart_data_start_col + 1,
                        min_row=chart_data_start_row,
                        max_row=chart_data_start_row + 3)
        cats = Reference(self.ws, min_col=chart_data_start_col,
                        min_row=chart_data_start_row + 1,
                        max_row=chart_data_start_row + 3)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # 添加到工作表
        self.ws.add_chart(chart, "P5")
        logger.info("生产完成度图表创建完成")

    def create_customer_chart(self):
        """创建客户分布饼图"""
        logger.info("创建客户分布图表...")

        # 使用现有的客户数据 (F9:F13)
        chart = PieChart()
        chart.title = "客户订单分布"
        chart.style = 10

        labels = Reference(self.ws, min_col=5, min_row=9, max_row=13)
        data = Reference(self.ws, min_col=6, min_row=8, max_row=13)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.dataLabels = openpyxl.chart.label.DataLabelList()
        chart.dataLabels.showPercent = True

        self.ws.add_chart(chart, "P15")
        logger.info("客户分布图表创建完成")

    def create_staffing_chart(self):
        """创建员工缺口对比图表"""
        logger.info("创建员工缺口图表...")

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "员工计划 vs 实际"
        chart.y_axis.title = "人数"

        # 数据范围：员工数据 L3:M8
        data1 = Reference(self.ws, min_col=12, min_row=2, max_row=8)  # Total Request
        data2 = Reference(self.ws, min_col=13, min_row=2, max_row=8)  # IN
        cats = Reference(self.ws, min_col=11, min_row=3, max_row=8)

        chart.add_data(data1, titles_from_data=True)
        chart.add_data(data2, titles_from_data=True)
        chart.set_categories(cats)

        self.ws.add_chart(chart, "V5")
        logger.info("员工缺口图表创建完成")

    def apply_conditional_formatting(self):
        """应用条件格式"""
        logger.info("应用条件格式...")

        # 规则1：完成率着色 (O7, O9, O11)
        logger.info("  应用规则1：完成率着色")
        rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=85, mid_color='FFEB84',
            end_type='num', end_value=100, end_color='63BE7B'
        )
        self.ws.conditional_formatting.add('P7:P11', rule)

        # 规则2：员工到岗率警告 (M3:M8)
        logger.info("  应用规则2：员工到岗率警告")
        red_fill = PatternFill(start_color='FFC7CE', fill_type='solid')
        red_font = Font(color='9C0006', bold=True)
        rule2 = CellIsRule(operator='lessThan', formula=['0.9*$L3'],
                          fill=red_fill, font=red_font)
        self.ws.conditional_formatting.add('M3:M8', rule2)

        # 规则3：产量数据条 (F8, C9, I8)
        logger.info("  应用规则3：产量数据条")
        from openpyxl.formatting.rule import DataBarRule
        data_bar = DataBarRule(start_type='num', start_value=0,
                              end_type='num', end_value=300,
                              color='4472C4')
        self.ws.conditional_formatting.add('F8', data_bar)
        self.ws.conditional_formatting.add('C9', data_bar)
        self.ws.conditional_formatting.add('I8', data_bar)

        logger.info("条件格式应用完成")

    def validate_formulas(self):
        """验证公式"""
        logger.info("验证公式...")

        error_count = 0
        for row in self.ws.iter_rows(min_row=2, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('='):
                        if '#REF!' in str(cell.value) or '#VALUE!' in str(cell.value):
                            logger.warning(f"  公式错误: {cell.coordinate} = {cell.value}")
                            error_count += 1

        if error_count == 0:
            logger.info("✓ 所有公式验证通过")
        else:
            logger.warning(f"✗ 发现 {error_count} 个公式错误")

        return error_count == 0

    def save_workbook(self):
        """保存工作簿"""
        logger.info(f"保存工作簿: {Config.OUTPUT_FILE}")
        self.wb.save(Config.OUTPUT_FILE)
        logger.info("工作簿保存成功")

    def run(self):
        """执行完整流程"""
        try:
            logger.info("=" * 70)
            logger.info("开始 Dashboard 增强流程")
            logger.info("=" * 70)

            # Step 1: 加载和备份
            self.load_workbook()
            self.create_backup(Config.BACKUP_FILE)

            # Step 2: 添加 KPI
            self.add_kpi_section()

            # Step 3: 创建图表
            self.create_production_chart()
            self.create_customer_chart()
            self.create_staffing_chart()

            # Step 4: 应用条件格式
            self.apply_conditional_formatting()

            # Step 5: 验证
            self.validate_formulas()

            # Step 6: 保存
            self.save_workbook()

            logger.info("=" * 70)
            logger.info("✅ Dashboard 增强流程完成成功")
            logger.info("=" * 70)

            return True

        except Exception as e:
            logger.error(f"错误: {e}", exc_info=True)
            logger.info("❌ Dashboard 增强流程失败")
            return False


# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数"""

    # 检查输入文件
    if not Config.INPUT_FILE.exists():
        logger.error(f"输入文件不存在: {Config.INPUT_FILE}")
        return False

    # 创建增强器并运行
    enhancer = DashboardEnhancer(Config.INPUT_FILE)
    success = enhancer.run()

    if success:
        logger.info("")
        logger.info("=" * 70)
        logger.info("最终摘要")
        logger.info("=" * 70)
        logger.info(f"输入文件: {Config.INPUT_FILE}")
        logger.info(f"输出文件: {Config.OUTPUT_FILE}")
        logger.info(f"备份文件: {Config.BACKUP_FILE}")
        logger.info("")
        logger.info("已添加功能:")
        logger.info("  ✓ 6 个 KPI 指标")
        logger.info("  ✓ 3 个可视化图表")
        logger.info("  ✓ 3 类条件格式规则")
        logger.info("")
        logger.info("=" * 70)

    return success


if __name__ == "__main__":
    import sys

    try:
        success = main()
        sys.exit(0 if success else 1)

    except KeyboardInterrupt:
        logger.info("用户中断")
        sys.exit(1)

    except Exception as e:
        logger.error(f"异常: {e}", exc_info=True)
        sys.exit(1)
