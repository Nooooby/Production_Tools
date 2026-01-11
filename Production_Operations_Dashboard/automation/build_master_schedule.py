"""
Master 排班汇总生成器

功能:
1. 规范 Master 输出字段: DeptID, Name, Station, StartTime, EndTime, Breaks
2. 汇总各部门排班表追加到 Master 表
3. 最终按 DeptID -> Station -> Name 排序
4. 可选保存副本 (导出)
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import logging
import shutil

import openpyxl
import pandas as pd


@dataclass(frozen=True)
class DepartmentConfig:
    dept_id: str
    sheet_name: str


class Config:
    """系统配置"""

    BASE_DIR = Path(r"C:\Projects\Production_management\Production_Operations_Dashboard")
    INPUT_FILE = BASE_DIR / "data" / "v39_Normalized_Colored.xlsx"
    OUTPUT_FILE = BASE_DIR / "data" / "v39_Normalized_Colored.xlsx"
    BACKUP_FILE = BASE_DIR / "data" / "v39_Normalized_Colored_backup_before_master.xlsx"
    LOG_DIR = BASE_DIR / "logs"

    MASTER_SHEET = "Master"

    # 部门配置 (按实际工作表名称调整)
    DEPARTMENTS = [
        DepartmentConfig(dept_id="Tray Pack", sheet_name="Tray Pack"),
        DepartmentConfig(dept_id="Cut-Up", sheet_name="Cut-Up"),
        DepartmentConfig(dept_id="Bagging", sheet_name="Bagging"),
    ]

    # 可选: 保存副本
    SAVE_COPY = True
    EXPORT_DIR = BASE_DIR / "exports"


OUTPUT_COLUMNS = ["DeptID", "Name", "Station", "StartTime", "EndTime", "Breaks"]
DATA_COLUMNS = ["Name", "Station", "StartTime", "EndTime", "Breaks"]


def setup_logging() -> logging.Logger:
    """设置日志系统"""
    Config.LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = Config.LOG_DIR / f"master_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    return logging.getLogger(__name__)


logger = setup_logging()


def build_department_df(source_path: Path, dept: DepartmentConfig) -> pd.DataFrame:
    """读取部门工作表并规范字段"""
    logger.info("读取部门工作表: %s", dept.sheet_name)
    df = pd.read_excel(source_path, sheet_name=dept.sheet_name, header=0)

    for column in DATA_COLUMNS:
        if column not in df.columns:
            logger.warning("部门 %s 缺少列 %s, 自动补空", dept.dept_id, column)
            df[column] = ""

    df["DeptID"] = dept.dept_id

    df = df[OUTPUT_COLUMNS]
    df = df.dropna(how="all", subset=DATA_COLUMNS)

    for column in DATA_COLUMNS:
        if df[column].dtype == "float64":
            df[column] = df[column].fillna("")

    return df


def combine_departments(source_path: Path) -> pd.DataFrame:
    """汇总所有部门数据"""
    frames = []
    for dept in Config.DEPARTMENTS:
        try:
            frames.append(build_department_df(source_path, dept))
        except ValueError as exc:
            logger.error("读取部门 %s 失败: %s", dept.sheet_name, exc)

    if not frames:
        logger.warning("未找到任何部门数据")
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    master_df = pd.concat(frames, ignore_index=True)

    def sort_key(series: pd.Series) -> pd.Series:
        return series.fillna("").astype(str).str.casefold()

    master_df = master_df.sort_values(
        by=["DeptID", "Station", "Name"],
        key=sort_key,
        kind="stable",
    )

    return master_df


def write_master_sheet(workbook_path: Path, master_df: pd.DataFrame) -> None:
    """写入 Master 工作表"""
    logger.info("写入 Master 工作表: %s", Config.MASTER_SHEET)
    wb = openpyxl.load_workbook(workbook_path)

    if Config.MASTER_SHEET in wb.sheetnames:
        del wb[Config.MASTER_SHEET]

    ws = wb.create_sheet(Config.MASTER_SHEET)
    ws.append(OUTPUT_COLUMNS)

    for row in master_df.itertuples(index=False):
        ws.append(list(row))

    wb.save(workbook_path)
    logger.info("Master 工作表写入完成")


def save_copy(source_path: Path) -> None:
    """保存副本 (导出)"""
    if not Config.SAVE_COPY:
        return

    Config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_path = Config.EXPORT_DIR / f"master_schedule_{timestamp}.xlsx"
    shutil.copy2(source_path, export_path)
    logger.info("导出副本: %s", export_path)


def create_backup() -> None:
    """创建备份"""
    if Config.BACKUP_FILE.exists():
        logger.info("备份已存在，跳过: %s", Config.BACKUP_FILE)
        return
    shutil.copy2(Config.INPUT_FILE, Config.BACKUP_FILE)
    logger.info("备份创建成功: %s", Config.BACKUP_FILE)


def main() -> None:
    """入口函数"""
    logger.info("开始生成 Master 排班汇总")

    if not Config.INPUT_FILE.exists():
        logger.error("未找到输入文件: %s", Config.INPUT_FILE)
        return

    create_backup()

    master_df = combine_departments(Config.INPUT_FILE)
    write_master_sheet(Config.OUTPUT_FILE, master_df)
    save_copy(Config.OUTPUT_FILE)

    logger.info("Master 排班汇总完成")


if __name__ == "__main__":
    main()
