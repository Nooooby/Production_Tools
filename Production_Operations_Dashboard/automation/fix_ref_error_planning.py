# -*- coding: utf-8 -*-
"""
fix_ref_error_planning.py - 修复 14_Production_Planning Row 6 的 #REF! 错误

问题：
- F6 和 G6 引用了不存在的列导致 #REF! 错误
- TrayPack 汇总行无法显示 Raw_kg 和 Cages

修复：
- F6: =SUMIF('06_Resource_Plan'!M:M,">0",'06_Resource_Plan'!L:L)
- G6: =SUMIF('06_Resource_Plan'!M:M,">0",'06_Resource_Plan'!M:M)

作者: Claude
日期: 2026-01-02
"""

import openpyxl
from datetime import datetime
import shutil
import os

# 配置
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

INPUT_FILE = os.path.join(PROJECT_ROOT, 'data', 'v39_Dashboard_Enhanced.xlsx')
OUTPUT_FILE = os.path.join(PROJECT_ROOT, 'data', 'v39_Dashboard_Enhanced.xlsx')
BACKUP_FILE = os.path.join(PROJECT_ROOT, 'data', f'v39_Dashboard_Enhanced_backup_ref_fix_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
LOG_FILE = os.path.join(PROJECT_ROOT, 'logs', f'fix_ref_error_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')

os.makedirs(os.path.join(PROJECT_ROOT, 'logs'), exist_ok=True)

def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_line + '\n')

def backup_file():
    """创建备份文件"""
    shutil.copy(INPUT_FILE, BACKUP_FILE)
    log(f"Backup created: {BACKUP_FILE}")

def check_resource_plan_structure(wb):
    """
    检查 06_Resource_Plan 的列结构
    确认 L 列和 M 列存在
    """
    ws = wb['06_Resource_Plan']

    log("Checking 06_Resource_Plan structure...")

    # 检查列头
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col_idx).value
        if cell_value:
            headers[col_idx] = cell_value
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            log(f"  Column {col_letter}: {cell_value}")

    # 查找关键列
    h_col = None
    l_col = None
    m_col = None

    for col_idx, header in headers.items():
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if 'Product_Group' in str(header) or header == 'H':
            h_col = col_letter
        elif 'Raw_kg' in str(header) or 'Raw' in str(header):
            l_col = col_letter
        elif 'Cages' in str(header) or header == 'Cages_Needed':
            m_col = col_letter

    log(f"Key columns found:")
    log(f"  Product_Group (H): {h_col if h_col else 'NOT FOUND'}")
    log(f"  Raw_kg_Needed (L): {l_col if l_col else 'NOT FOUND'}")
    log(f"  Cages_Needed (M): {m_col if m_col else 'NOT FOUND'}")

    return h_col, l_col, m_col

def get_current_f6_g6_formulas(wb):
    """获取当前 F6 和 G6 的公式"""
    ws = wb['14_Production_Planning']

    f6_formula = ws['F6'].value
    g6_formula = ws['G6'].value

    log("Current formulas:")
    log(f"  F6: {f6_formula}")
    log(f"  G6: {g6_formula}")

    # 检查是否有 #REF! 错误
    has_ref_error = '#REF!' in str(f6_formula) or '#REF!' in str(g6_formula)

    if has_ref_error:
        log("  [ERROR] #REF! error detected!")
    else:
        log("  [OK] No #REF! errors found")

    return f6_formula, g6_formula, has_ref_error

def fix_f6_g6_formulas(wb, l_col, m_col):
    """
    修复 F6 和 G6 的公式

    F6: 汇总所有有 Cages 需求的产品的 Raw_kg
    G6: 汇总所有有 Cages 需求的产品的 Cages
    """
    ws = wb['14_Production_Planning']

    log("Fixing F6 and G6 formulas...")

    # 如果找不到 L 和 M 列，使用默认值
    if not l_col:
        l_col = 'L'
        log(f"  Warning: Raw_kg column not found, using default 'L'")

    if not m_col:
        m_col = 'M'
        log(f"  Warning: Cages column not found, using default 'M'")

    # 修复 F6: 汇总 Raw_kg
    old_f6 = ws['F6'].value
    ws['F6'] = f"=SUMIF('06_Resource_Plan'!{m_col}:{m_col},\">0\",'06_Resource_Plan'!{l_col}:{l_col})"
    new_f6 = ws['F6'].value
    log(f"  F6 updated:")
    log(f"    Old: {old_f6}")
    log(f"    New: {new_f6}")

    # 修复 G6: 汇总 Cages
    old_g6 = ws['G6'].value
    ws['G6'] = f"=SUMIF('06_Resource_Plan'!{m_col}:{m_col},\">0\",'06_Resource_Plan'!{m_col}:{m_col})"
    new_g6 = ws['G6'].value
    log(f"  G6 updated:")
    log(f"    Old: {old_g6}")
    log(f"    New: {new_g6}")

    return True

def verify_fix(wb):
    """验证修复是否成功"""
    ws = wb['14_Production_Planning']

    log("Verifying fix...")

    f6_formula = ws['F6'].value
    g6_formula = ws['G6'].value

    # 检查是否还有 #REF! 错误
    has_ref_error = False
    if '#REF!' in str(f6_formula):
        log("  [ERROR] F6 still has #REF! error")
        has_ref_error = True
    else:
        log(f"  [OK] F6 formula OK: {f6_formula}")

    if '#REF!' in str(g6_formula):
        log("  [ERROR] G6 still has #REF! error")
        has_ref_error = True
    else:
        log(f"  [OK] G6 formula OK: {g6_formula}")

    # 检查其他关键单元格
    g9_formula = ws['G9'].value
    e34_formula = ws['E34'].value

    log(f"  Related formulas:")
    log(f"    G9: {g9_formula}")
    log(f"    E34: {e34_formula}")

    if has_ref_error:
        log("[FAILED] Verification FAILED - #REF! errors still exist")
        return False
    else:
        log("[PASSED] Verification PASSED - All formulas OK")
        return True

def main():
    log("=" * 60)
    log("Fix #REF! Error in 14_Production_Planning - Start")
    log("=" * 60)

    # 创建备份
    backup_file()

    # 加载工作簿
    log(f"Loading workbook: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    # 1. 检查 06_Resource_Plan 结构
    h_col, l_col, m_col = check_resource_plan_structure(wb)

    # 2. 获取当前 F6 和 G6 公式
    old_f6, old_g6, has_error = get_current_f6_g6_formulas(wb)

    if not has_error:
        log("No #REF! errors found. Nothing to fix.")
        wb.close()
        return True

    # 3. 修复 F6 和 G6
    fix_f6_g6_formulas(wb, l_col, m_col)

    # 4. 验证修复
    success = verify_fix(wb)

    if not success:
        log("Fix failed. Not saving changes.")
        wb.close()
        return False

    # 5. 保存
    log(f"Saving workbook: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    wb.close()

    log("=" * 60)
    log("Fix #REF! Error - Complete")
    log(f"Backup file: {BACKUP_FILE}")
    log(f"Log file: {LOG_FILE}")
    log("=" * 60)

    return True

if __name__ == '__main__':
    success = main()
    exit(0 if success else 1)
