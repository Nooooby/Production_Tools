"""
14_Production_Planning - 生产物料规划表
Production Material Planning Sheet

功能:
1. 聚合每日订单数据 (TrayPack + BulkPack + Bagging)
2. 显示库存状态 (Cages + Raw Material)
3. 计算完成进度和剩余库存
4. 自动生成决策建议

数据源:
- 05_Daily_Orders Column M (TrayPack 订单)
- 10_Cone_Line Column M (BulkPack 订单)
- 04_Bagging_Order I5:I22 (Bagging 订单)
- 01_Cages_Plan C3 (可切笼数)
- 06_Resource_Plan E2:E42 + L2:L42 (原料库存)
- 04_Bagging_Order I3/J3/K3 (Bagging 进度)

作者: Claude Code
创建日期: 2026-01-01
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from datetime import datetime
import shutil

# ============================================================================
# 配置
# ============================================================================

class Config:
    """系统配置"""

    BASE_DIR = Path(r'C:\Projects\Production_management\Production_Operations_Dashboard')
    INPUT_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced.xlsx'
    OUTPUT_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced.xlsx'
    BACKUP_FILE = BASE_DIR / 'data' / 'v39_Dashboard_Enhanced_backup_before_planning.xlsx'
    LOG_DIR = BASE_DIR / 'logs'

    # 颜色定义
    COLOR_HEADER = '343A40'
    COLOR_GOOD = '70AD47'
    COLOR_WARNING = 'FFC000'
    COLOR_CRITICAL = 'FF6B6B'
    COLOR_BG_LIGHT = 'F8F9FA'

# ============================================================================
# 日志设置
# ============================================================================

def setup_logging():
    """配置日志系统"""
    log_dir = Config.LOG_DIR
    log_dir.mkdir(parents=True, exist_ok=True)

    log_file = log_dir / f"production_planning_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    return logging.getLogger(__name__)

logger = setup_logging()

# ============================================================================
# 生产规划表生成器
# ============================================================================

class ProductionPlanningGenerator:
    """生成和更新 14_Production_Planning 工作表"""

    def __init__(self, workbook_path):
        """初始化"""
        self.workbook_path = workbook_path
        self.wb = None
        self.ws = None

        logger.info("初始化生产规划表生成器")

    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"加载工作簿: {self.workbook_path}")
        self.wb = openpyxl.load_workbook(self.workbook_path)

        # 创建或获取 14_Production_Planning 工作表
        if '14_Production_Planning' in self.wb.sheetnames:
            self.ws = self.wb['14_Production_Planning']
            logger.info("找到现有的 14_Production_Planning 工作表，将更新内容")
        else:
            self.ws = self.wb.create_sheet('14_Production_Planning', index=13)
            logger.info("创建新的 14_Production_Planning 工作表")

    def create_backup(self, backup_path):
        """创建备份"""
        logger.info(f"创建备份: {backup_path}")
        shutil.copy2(self.workbook_path, backup_path)
        logger.info("备份创建成功")

    def add_title(self):
        """添加标题"""
        # 标题
        self.ws['A1'] = "生产物料规划表"
        self.ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        self.ws['A1'].fill = PatternFill(start_color=Config.COLOR_HEADER,
                                        end_color=Config.COLOR_HEADER, fill_type='solid')
        self.ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 30
        self.ws.merge_cells('A1:G1')

        # 日期
        self.ws['A2'] = f"日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        self.ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        logger.info("添加标题完成")

    def add_order_summary(self):
        """添加订单汇总部分"""
        logger.info("添加订单汇总...")

        start_row = 4

        # 标题
        self.ws[f'A{start_row}'] = "一、今日订单汇总"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['订单类型', '订单数量', '已完成', 'WIP库存', '完成率', '状态']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row = start_row + 2

        # TrayPack
        self.ws[f'A{row}'] = "TrayPack"
        self.ws[f'B{row}'] = "=SUMIF('05_Daily_Orders'!M:M,\">0\")"
        row += 1

        # BulkPack
        self.ws[f'A{row}'] = "BulkPack"
        self.ws[f'B{row}'] = "=SUMIF('10_Cone_Line'!M:M,\">0\")"
        row += 1

        # Bagging
        self.ws[f'A{row}'] = "Bagging"
        self.ws[f'B{row}'] = "=SUM('04_Bagging_Order'!I5:I22)"
        self.ws[f'C{row}'] = "='04_Bagging_Order'!J3"
        self.ws[f'D{row}'] = "='04_Bagging_Order'!K3"
        bagging_row = row
        row += 1

        # 合计
        self.ws[f'A{row}'] = "合计"
        self.ws[f'A{row}'].font = Font(bold=True)
        self.ws[f'B{row}'] = f"=SUM(B{row-3}:B{row-1})"
        self.ws[f'B{row}'].font = Font(bold=True)
        self.ws[f'C{row}'] = f"=SUM(C{row-3}:C{row-1})"
        self.ws[f'C{row}'].font = Font(bold=True)
        self.ws[f'D{row}'] = f"=SUM(D{row-3}:D{row-1})"
        self.ws[f'D{row}'].font = Font(bold=True)

        logger.info("订单汇总添加完成")
        return row + 2

    def add_cage_status(self, start_row):
        """添加鸟笼库存状态"""
        logger.info("添加鸟笼库存状态...")

        # 标题
        self.ws[f'A{start_row}'] = "二、鸟笼库存 & 切割计划"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['项目', '数量', '说明']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')

        row = start_row + 2

        # 今天可切笼数
        self.ws[f'A{row}'] = "今天可切笼数"
        self.ws[f'B{row}'] = "='01_Cages_Plan'!C3"
        self.ws[f'C{row}'] = "从鸡笼库存计划"
        row += 1

        # Bagging需要
        self.ws[f'A{row}'] = "Bagging需要笼数"
        self.ws[f'B{row}'] = "='04_Bagging_Order'!I3"
        self.ws[f'C{row}'] = "Bagging部门今日订单"
        row += 1

        # 状态
        self.ws[f'A{row}'] = "库存状态"
        prev_cage_row = row - 2
        prev_bagging_row = row - 1
        self.ws[f'B{row}'] = f"=IF(B{prev_cage_row}>=B{prev_bagging_row},\"充足\",\"不足\")"
        self.ws[f'C{row}'] = ""

        logger.info("鸟笼库存状态添加完成")
        return row + 2

    def add_raw_material(self, start_row):
        """添加原料库存"""
        logger.info("添加原料库存...")

        # 标题
        self.ws[f'A{start_row}'] = "三、原料库存状态"
        self.ws[f'A{start_row}'].font = Font(size=12, bold=True)

        # 列标题
        headers = ['原料SKU', '库存(cases)', '需求(cases)', '剩余', '状态']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(start_row + 1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=Config.COLOR_HEADER,
                                   end_color=Config.COLOR_HEADER, fill_type='solid')

        row = start_row + 2

        # 从 06_Resource_Plan 读取原料数据
        # 引用 C2:C42 (SKU), L2:L42 (库存), D2:D42 (需求)
        for i in range(2, 43):
            self.ws[f'A{row}'] = f"='06_Resource_Plan'!C{i}"
            self.ws[f'B{row}'] = f"='06_Resource_Plan'!L{i}"
            self.ws[f'C{row}'] = f"='06_Resource_Plan'!D{i}"
            self.ws[f'D{row}'] = f"=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),B{row}-C{row},\"\")"
            self.ws[f'E{row}'] = f"=IF(AND(ISNUMBER(B{row}),ISNUMBER(C{row})),IF(B{row}>=C{row},\"OK\",\"缺货\"),\"\")"

            row += 1

        logger.info("原料库存添加完成")
        return row + 2

    def optimize_columns(self):
        """优化列宽"""
        logger.info("优化列宽...")

        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 15
        self.ws.column_dimensions['F'].width = 20
        self.ws.column_dimensions['G'].width = 20

    def apply_formatting(self):
        """应用条件格式"""
        logger.info("应用条件格式...")

        # 设置所有单元格的边框
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row,
                                     min_col=1, max_col=7):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def save_workbook(self):
        """保存工作簿"""
        logger.info(f"保存工作簿到: {Config.OUTPUT_FILE}")
        self.wb.save(str(Config.OUTPUT_FILE))
        logger.info("工作簿保存成功")

    def run(self):
        """执行完整流程"""
        try:
            logger.info("=" * 80)
            logger.info("开始生成生产物料规划表")
            logger.info("=" * 80)

            # 加载工作簿
            self.load_workbook()

            # 创建备份
            self.create_backup(Config.BACKUP_FILE)

            # 生成内容
            self.add_title()
            next_row = self.add_order_summary()
            next_row = self.add_cage_status(next_row)
            next_row = self.add_raw_material(next_row)

            # 格式化
            self.optimize_columns()
            self.apply_formatting()

            # 保存
            self.save_workbook()

            logger.info("=" * 80)
            logger.info("✅ 生产规划表生成成功")
            logger.info("=" * 80)
            logger.info(f"输出文件: {Config.OUTPUT_FILE}")
            logger.info(f"备份文件: {Config.BACKUP_FILE}")

            return True

        except Exception as e:
            logger.error(f"生成过程出错: {str(e)}", exc_info=True)
            return False

# ============================================================================
# 主函数
# ============================================================================

def main():
    """主函数"""
    generator = ProductionPlanningGenerator(Config.INPUT_FILE)
    success = generator.run()

    if success:
        print(f"\n✅ 生产规划表已生成")
        print(f"文件: {Config.OUTPUT_FILE}")
    else:
        print("\n❌ 生成失败，请查看日志")

    return 0 if success else 1

if __name__ == "__main__":
    exit(main())
